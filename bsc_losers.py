#!/usr/bin/env python3
"""
BSC Token Loss Finder
─────────────────────
通过 BscScan V2 API（免费）获取完整历史 Swap 事件，计算每个地址的 BNB 盈亏。

【获取免费 API Key】
  1. 注册 https://etherscan.io/register（免费）
  2. 进入 Profile → API Keys → Create
  3. 将 Key 传给 --key 参数（同时支持 BSC + ETH + Base 等全链）

用法:
  python3 bsc_losers.py --token 0xABC... --key YOUR_KEY
  python3 bsc_losers.py --pair  0xXYZ... --key YOUR_KEY --top 30
  python3 bsc_losers.py --pair  0xXYZ... --key YOUR_KEY --out losers.xlsx

注: 无 API key 时使用演示模式生成示例数据
"""

import argparse, sys, time, math, requests, openpyxl, random
from openpyxl.styles import Font, PatternFill, Alignment
from collections import defaultdict
from datetime import datetime

# ───── 常量 ──────────────────────────────────────────
WBNB       = "0xbb4cdb9cbd36b01bd1cbaebf2de08d9173bc095c"
FACTORY_V2 = "0xcA143Ce32Fe78f1f7019d7d551a6402fC5350c73"
SWAP_TOPIC = "0xd78ad95fa46c994b6551d0da85fc275fe613ce37657fb8d5e3d130840159d822"
WEI        = 10 ** 18
BSC_CHAIN  = "56"

BSCSCAN_V2 = "https://api.etherscan.io/v2/api"
BSC_RPC    = "https://bsc.publicnode.com"   # 只用于最新块 / token/pair 查询


# ───── RPC 工具（仅用于最新块和合约调用）────────────
def rpc(method, params, timeout=15):
    r = requests.post(BSC_RPC, json={
        "jsonrpc":"2.0","method":method,"params":params,"id":1
    }, timeout=timeout)
    return r.json().get("result")


def current_block():
    return int(rpc("eth_blockNumber", []), 16)


def eth_call(to, data):
    return rpc("eth_call", [{"to": to, "data": data}, "latest"])


# ───── PancakeSwap V2 ─────────────────────────────────
def find_pair(token: str) -> str:
    """从 PancakeSwap Factory 获取 token/WBNB pair 地址"""
    ta = token.lower().replace("0x","").zfill(64)
    tb = WBNB.replace("0x","").zfill(64)
    sel = "0xe6a43905"
    for args in [ta+tb, tb+ta]:
        res = eth_call(FACTORY_V2, sel+args)
        if res:
            addr = "0x" + res[-40:]
            if addr.lower() != "0x"+"0"*40:
                return addr
    return None


def pair_tokens(pair: str):
    t0 = eth_call(pair, "0x0dfe1681")
    t1 = eth_call(pair, "0xd21220a7")
    return (
        ("0x" + (t0 or "0"*64)[-40:]).lower(),
        ("0x" + (t1 or "0"*64)[-40:]).lower(),
    )


# ───── BscScan V2 getLogs ─────────────────────────────
def bscscan_get_logs(pair: str, api_key: str, page: int = 1, offset: int = 1000):
    """通过 BscScan V2 API 获取 Swap 事件（支持完整历史）"""
    params = {
        "chainid":   BSC_CHAIN,
        "module":    "logs",
        "action":    "getLogs",
        "address":   pair,
        "topic0":    SWAP_TOPIC,
        "fromBlock": "0",
        "toBlock":   "latest",
        "page":      page,
        "offset":    offset,
        "apikey":    api_key,
    }
    r = requests.get(BSCSCAN_V2, params=params, timeout=30)
    d = r.json()
    status  = d.get("status")
    message = d.get("message","")
    result  = d.get("result",[])

    if status == "1":
        return result, None
    elif status == "0" and message == "No records found":
        return [], None
    else:
        return [], message


def fetch_all_logs(pair: str, api_key: str):
    """分页获取全部 Swap 日志"""
    all_logs = []
    page = 1
    print("  分页获取 Swap 日志：", end="", flush=True)
    while True:
        logs, err = bscscan_get_logs(pair, api_key, page=page, offset=1000)
        if err:
            print(f"\n  ⚠️  BscScan 错误: {err}")
            break
        all_logs.extend(logs)
        print(f" p{page}({len(logs)})", end="", flush=True)
        if len(logs) < 1000:
            break   # 最后一页
        page += 1
        time.sleep(0.25)   # 免费 API 5 req/s
    print(f" → 共 {len(all_logs)} 条\n")
    return all_logs


# ───── P&L 计算 ───────────────────────────────────────
def compute_pnl(logs: list, bnb_is_token1: bool):
    pnl      = defaultdict(float)
    buys     = defaultdict(int)
    sells    = defaultdict(int)
    buy_vol  = defaultdict(float)
    sell_vol = defaultdict(float)
    txs      = defaultdict(set)

    for log in logs:
        data = log.get("data","").replace("0x","")
        if len(data) < 256: continue

        a0in  = int(data[0:64],    16)
        a1in  = int(data[64:128],  16)
        a0out = int(data[128:192], 16)
        a1out = int(data[192:256], 16)

        topics = log.get("topics", [])
        to_addr = ("0x" + topics[2][-40:]).lower() if len(topics) > 2 else None
        if not to_addr: continue

        bnb_in  = (a1in  if bnb_is_token1 else a0in)  / WEI
        bnb_out = (a1out if bnb_is_token1 else a0out) / WEI
        tx      = log.get("transactionHash","")

        if bnb_in  > 0: pnl[to_addr] -= bnb_in;  buys[to_addr]  += 1; buy_vol[to_addr]  += bnb_in
        if bnb_out > 0: pnl[to_addr] += bnb_out; sells[to_addr] += 1; sell_vol[to_addr] += bnb_out
        txs[to_addr].add(tx)

    return [{
        "address":      addr,
        "net_bnb":      round(net, 8),
        "buy_count":    buys[addr],
        "sell_count":   sells[addr],
        "buy_vol_bnb":  round(buy_vol[addr], 6),
        "sell_vol_bnb": round(sell_vol[addr], 6),
        "tx_count":     len(txs[addr]),
    } for addr, net in pnl.items()]


# ───── 演示模式 ───────────────────────────────────────
def demo_mode(pair: str):
    """无 API key 时生成示例数据，说明脚本功能"""
    print("\n⚠️  演示模式（无真实数据）—— 传入 --key 获得真实结果\n")
    rng = random.Random(42)
    traders = []
    for i in range(80):
        buys    = rng.randint(1, 10)
        sells   = rng.randint(0, buys)
        bv      = round(rng.uniform(0.05, 3.0), 4)
        sv      = round(bv * rng.uniform(0.3, 1.4), 4)
        net     = round(sv - bv, 6)
        addr    = "0x" + "".join(rng.choices("0123456789abcdef", k=40))
        traders.append({"address":addr,"net_bnb":net,"buy_count":buys,
                        "sell_count":sells,"buy_vol_bnb":bv,"sell_vol_bnb":sv,"tx_count":buys+sells})
    return traders


# ───── Excel 导出 ─────────────────────────────────────
def export_excel(losers, all_traders, token, pair, out_path):
    wb = openpyxl.Workbook()

    # Sheet 1: 亏损地址排行
    ws = wb.active; ws.title = "亏损排行"
    RED  = PatternFill("solid", fgColor="C0392B")
    PINK = PatternFill("solid", fgColor="FDECEA")
    HF   = Font(bold=True, color="FFFFFF", size=11)
    CT   = Alignment(horizontal="center")
    hdrs = ["排名","钱包地址","净亏损 (BNB)","买入次数","卖出次数","买入量 (BNB)","卖出量 (BNB)","交易次数"]
    wids = [6, 46, 16, 10, 10, 16, 16, 10]
    for c, (h, w) in enumerate(zip(hdrs, wids), 1):
        cell = ws.cell(1, c, h); cell.fill=RED; cell.font=HF; cell.alignment=CT
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    for rank, t in enumerate(losers, 1):
        for c, v in enumerate([rank, t["address"], t["net_bnb"], t["buy_count"],
                                t["sell_count"], t["buy_vol_bnb"], t["sell_vol_bnb"], t["tx_count"]], 1):
            cell = ws.cell(rank+1, c, v); cell.fill = PINK
            if c == 3 and isinstance(v, float): cell.number_format = "0.000000"

    # Sheet 2: 全部交易者（含盈利）
    ws3 = wb.create_sheet("全部交易者")
    all_sorted = sorted(all_traders, key=lambda x: x["net_bnb"])
    GREEN = PatternFill("solid", fgColor="E9F7EF")
    hdrs2 = ["排名","钱包地址","净盈亏 (BNB)","买入次数","卖出次数","买入量 (BNB)","卖出量 (BNB)","交易次数"]
    for c, (h, w) in enumerate(zip(hdrs2, wids), 1):
        cell = ws3.cell(1, c, h); cell.fill=RED; cell.font=HF; cell.alignment=CT
        ws3.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    for rank, t in enumerate(all_sorted, 1):
        fill = PINK if t["net_bnb"] < 0 else GREEN
        for c, v in enumerate([rank, t["address"], t["net_bnb"], t["buy_count"],
                                t["sell_count"], t["buy_vol_bnb"], t["sell_vol_bnb"], t["tx_count"]], 1):
            cell = ws3.cell(rank+1, c, v); cell.fill = fill

    # Sheet 3: 汇总
    ws2 = wb.create_sheet("汇总")
    tl = sum(t["net_bnb"] for t in losers)
    tp = sum(t["net_bnb"] for t in all_traders if t["net_bnb"] >= 0)
    summary = [
        ("代币地址", token or "（未指定）"),
        ("Pair 地址", pair),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("数据来源", "BscScan V2 API (Etherscan)"),
        ("─"*16, "─"*30),
        ("总交易地址", len(all_traders)),
        ("亏损地址", len(losers)),
        ("盈利地址", len(all_traders)-len(losers)),
        ("亏损占比", f"{len(losers)/len(all_traders)*100:.1f}%" if all_traders else "N/A"),
        ("─"*16, "─"*30),
        ("总亏损 (BNB)", round(tl, 6)),
        ("总盈利 (BNB)", round(tp, 6)),
        ("平均亏损 (BNB)", round(tl/len(losers), 6) if losers else 0),
        ("最大单人亏损 (BNB)", losers[0]["net_bnb"] if losers else 0),
        ("最大亏损地址", losers[0]["address"] if losers else ""),
    ]
    ws2.column_dimensions["A"].width = 22; ws2.column_dimensions["B"].width = 50
    for i, (k, v) in enumerate(summary, 1):
        ka = ws2.cell(i,1,k); vb = ws2.cell(i,2,v)
        if not k.startswith("─"): ka.font = Font(bold=True)

    wb.save(out_path)
    return out_path


# ───── 主程序 ─────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(
        description="BSC Token Loss Finder — 找出持仓亏损的钱包地址",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python3 bsc_losers.py --pair 0xXYZ --key YOUR_ETHERSCAN_KEY
  python3 bsc_losers.py --token 0xABC --key YOUR_KEY --top 50

免费 API Key: https://etherscan.io/register → Profile → API Keys
        """
    )
    ap.add_argument("--token", help="代币合约地址（自动查找 WBNB pair）")
    ap.add_argument("--pair",  help="PancakeSwap V2 Pair 地址")
    ap.add_argument("--key",   default="", help="Etherscan API Key（免费）")
    ap.add_argument("--top",   type=int, default=20, help="显示亏损前 N 名（默认20）")
    ap.add_argument("--out",   default="", help="Excel 输出路径")
    ap.add_argument("--demo",  action="store_true", help="演示模式（生成示例数据）")
    args = ap.parse_args()

    if not args.token and not args.pair and not args.demo:
        ap.error("请提供 --token 或 --pair（或使用 --demo 体验演示）")

    print("\n" + "="*57)
    print("  💸  BSC Token Loss Finder  |  On-chain Analysis")
    print("="*57 + "\n")

    # 1. 确定 pair 地址
    token_addr = (args.token or "").lower()
    pair_addr  = (args.pair  or "").lower()

    if not args.demo:
        if not pair_addr and token_addr:
            print(f"🔗 查找 {token_addr[:12]}... 的 WBNB 交易对...")
            pair_addr = find_pair(token_addr)
            if not pair_addr:
                print("❌ 未找到 PancakeSwap V2 WBNB 交易对")
                sys.exit(1)
            print(f"✅ Pair: {pair_addr}")

        # 2. 确认 WBNB 位置
        token0, token1 = pair_tokens(pair_addr)
        bnb_is_token1 = (token1 == WBNB)
        print(f"   token0: {token0}")
        print(f"   token1: {token1}")
        print(f"   WBNB = {'token1' if bnb_is_token1 else 'token0'}\n")

        # 3. 获取 Swap 数据
        if not args.key:
            print("⚠️  未提供 --key，进入演示模式\n")
            print("   获取免费 Key: https://etherscan.io/register")
            print("   注册后 → Profile → API Keys → Create\n")
            all_traders = demo_mode(pair_addr)
            pair_addr = pair_addr or "0xDEMO"
        else:
            print(f"📡 获取 Swap 历史（BscScan V2 API）...")
            logs = fetch_all_logs(pair_addr, args.key)
            if not logs:
                print("⚠️  未找到 Swap 记录（可能: pair 错误 / 没有历史交易）")
                sys.exit(0)
            print(f"🔢 计算各地址盈亏...")
            all_traders = compute_pnl(logs, bnb_is_token1)
    else:
        all_traders = demo_mode("0xDEMO")
        pair_addr = "0xDEMO"
        token_addr = "0xDEMO"
        bnb_is_token1 = True

    # 4. 统计
    losers  = sorted([t for t in all_traders if t["net_bnb"] < 0], key=lambda x: x["net_bnb"])
    winners = [t for t in all_traders if t["net_bnb"] >= 0]
    tl = sum(t["net_bnb"] for t in losers)
    tp = sum(t["net_bnb"] for t in winners)

    print(f"\n{'='*57}")
    print(f"  📊  统计结果")
    print(f"{'='*57}")
    print(f"  总交易地址:   {len(all_traders):>8,}")
    print(f"  亏损人数:     {len(losers):>8,}   ({len(losers)/len(all_traders)*100:.1f}%)")
    print(f"  盈利人数:     {len(winners):>8,}   ({len(winners)/len(all_traders)*100:.1f}%)")
    print(f"  总亏损:       {tl:>12.4f} BNB")
    print(f"  总盈利:       {tp:>12.4f} BNB")
    if losers:
        print(f"  平均亏损:     {tl/len(losers):>12.4f} BNB")
        print(f"  最大亏损:     {losers[0]['net_bnb']:>12.4f} BNB  ({losers[0]['address'][:12]}...)")

    top_n = min(args.top, len(losers))
    if top_n:
        print(f"\n💸 亏损 Top {top_n}:")
        print(f"  {'排名':>4}  {'地址':<44}  {'亏损(BNB)':>12}  {'买/卖':>6}")
        print("  " + "─"*72)
        for i, t in enumerate(losers[:top_n], 1):
            print(f"  {i:>4}  {t['address']:<44}  {t['net_bnb']:>12.6f}  {t['buy_count']}/{t['sell_count']}")
    else:
        print("\n✅ 无亏损地址（所有交易者均盈利或数据为演示）")

    # 5. Excel
    out = args.out or f"bsc_losers_{pair_addr[:8]}.xlsx"
    export_excel(losers, all_traders, token_addr or pair_addr, pair_addr, out)
    print(f"\n✅ Excel 已导出: {out}")
    print(f"   包含 3 个 Sheet: 亏损排行 / 全部交易者 / 汇总\n")


if __name__ == "__main__":
    main()
