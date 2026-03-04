#!/usr/bin/env python3
"""
Worst Traders Scraper
─────────────────────
从 GMGN 抓取某 Token 的 Top 100 交易员数据，
筛选出所有亏损地址并生成 Excel 报告。

用法:
  python3 worst_traders.py --token 0xABC... --chain bsc --apify-key YOUR_KEY
  python3 worst_traders.py --token 0xABC... --chain bsc --apify-key YOUR_KEY --demo
"""

import argparse
import json
import sys
import time
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from datetime import datetime

APIFY_BASE = "https://api.apify.com/v2"
ACTOR_ID = "muhammetakkurtt~gmgn-token-traders-scraper"

CHAINS = ["bsc", "eth", "sol", "base", "tron"]


# ──────────────────────────────
# Apify 调用
# ──────────────────────────────

def run_actor(token_address: str, chain: str, apify_key: str) -> list:
    """运行 Apify Actor 并等待结果"""
    print(f"🚀 启动 Apify Actor: {ACTOR_ID}")
    print(f"   Token: {token_address}")
    print(f"   Chain: {chain.upper()}")

    # 启动 Actor run
    run_url = f"{APIFY_BASE}/acts/{ACTOR_ID}/runs?token={apify_key}"
    payload = {
        "tokenAddress": token_address,
        "chain": chain,
        "limit": 100,
    }
    r = requests.post(run_url, json=payload, timeout=30)
    r.raise_for_status()
    run_data = r.json()["data"]
    run_id = run_data["id"]
    print(f"   Run ID: {run_id}")

    # 等待完成
    print("⏳ 等待 Actor 完成...")
    for attempt in range(60):  # 最多等 5 分钟
        time.sleep(5)
        status_url = f"{APIFY_BASE}/actor-runs/{run_id}?token={apify_key}"
        sr = requests.get(status_url, timeout=15).json()["data"]
        status = sr.get("status", "")
        print(f"   [{attempt+1}] 状态: {status}")
        if status == "SUCCEEDED":
            break
        if status in ("FAILED", "ABORTED", "TIMED-OUT"):
            raise RuntimeError(f"Actor run {status}: {sr.get('statusMessage', '')}")
    else:
        raise TimeoutError("Actor 超时（5分钟未完成）")

    # 获取数据集
    dataset_id = sr["defaultDatasetId"]
    print(f"✅ 完成！Dataset: {dataset_id}")
    items_url = f"{APIFY_BASE}/datasets/{dataset_id}/items?token={apify_key}&clean=true&format=json"
    ir = requests.get(items_url, timeout=30)
    ir.raise_for_status()
    return ir.json()


# ──────────────────────────────
# Demo 数据（测试用）
# ──────────────────────────────

def demo_data(token_address: str, chain: str) -> list:
    """生成模拟数据用于演示"""
    import random
    random.seed(42)
    traders = []
    for i in range(100):
        addr = "0x" + "".join(random.choices("0123456789abcdef", k=40))
        profit = round(random.uniform(-5000, 3000), 2)
        buys = random.randint(1, 20)
        sells = random.randint(0, buys)
        traders.append({
            "wallet_address": addr,
            "realized_profit": profit,
            "unrealized_profit": round(random.uniform(-500, 500), 2),
            "total_profit": profit,
            "buy_30d": buys,
            "sell_30d": sells,
            "profit_change": round(random.uniform(-90, 200), 1),
            "win_rate": round(random.uniform(0, 1), 2) if sells > 0 else 0,
            "buy_volume_30d": round(random.uniform(100, 50000), 2),
            "sell_volume_30d": round(random.uniform(0, 50000), 2),
        })
    return traders


# ──────────────────────────────
# 数据解析
# ──────────────────────────────

def extract_profit(item: dict) -> float:
    """从 GMGN 数据中提取总盈亏（USD）"""
    # GMGN 可能返回的字段名变体
    for key in ["realized_profit", "total_profit_usd", "profit", "pnl", "realizedProfit"]:
        if key in item and item[key] is not None:
            try:
                return float(item[key])
            except (ValueError, TypeError):
                pass
    return 0.0


def parse_traders(raw_items: list) -> list:
    """标准化数据结构"""
    traders = []
    for item in raw_items:
        profit = extract_profit(item)
        traders.append({
            "wallet": item.get("wallet_address") or item.get("address") or item.get("wallet") or "unknown",
            "realized_profit_usd": profit,
            "unrealized_profit_usd": float(item.get("unrealized_profit") or 0),
            "total_profit_usd": profit + float(item.get("unrealized_profit") or 0),
            "buy_count": int(item.get("buy_30d") or item.get("buys") or 0),
            "sell_count": int(item.get("sell_30d") or item.get("sells") or 0),
            "win_rate": float(item.get("win_rate") or 0),
            "buy_volume_usd": float(item.get("buy_volume_30d") or item.get("buyVolume") or 0),
            "sell_volume_usd": float(item.get("sell_volume_30d") or item.get("sellVolume") or 0),
            "profit_pct": float(item.get("profit_change") or 0),
        })
    return traders


# ──────────────────────────────
# Excel 导出
# ──────────────────────────────

def export_excel(losers: list, token_address: str, chain: str, all_count: int) -> str:
    wb = openpyxl.Workbook()

    # ── Sheet 1: 亏损用户明细 ──
    ws = wb.active
    ws.title = "Worst Traders"

    RED    = PatternFill("solid", fgColor="C0392B")
    LIGHT  = PatternFill("solid", fgColor="FDECEA")
    HEADER = Font(bold=True, color="FFFFFF", size=11)
    CENTER = Alignment(horizontal="center")

    headers = [
        "排名", "钱包地址", "已实现亏损 (USD)",
        "未实现盈亏 (USD)", "总亏损 (USD)",
        "买入次数", "卖出次数", "胜率", "买入量 (USD)", "卖出量 (USD)"
    ]
    col_widths = [6, 46, 20, 20, 18, 10, 10, 10, 16, 16]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(1, col, h)
        cell.fill = RED
        cell.font = HEADER
        cell.alignment = CENTER
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    for rank, t in enumerate(losers, 1):
        row = rank + 1
        ws.cell(row, 1, rank)
        ws.cell(row, 2, t["wallet"])
        ws.cell(row, 3, round(t["realized_profit_usd"], 2))
        ws.cell(row, 4, round(t["unrealized_profit_usd"], 2))
        ws.cell(row, 5, round(t["total_profit_usd"], 2))
        ws.cell(row, 6, t["buy_count"])
        ws.cell(row, 7, t["sell_count"])
        ws.cell(row, 8, f"{t['win_rate']*100:.1f}%")
        ws.cell(row, 9, round(t["buy_volume_usd"], 2))
        ws.cell(row, 10, round(t["sell_volume_usd"], 2))
        for col in range(1, 11):
            ws.cell(row, col).fill = LIGHT

    # ── Sheet 2: 汇总 ──
    ws2 = wb.create_sheet("汇总")
    BLUE = PatternFill("solid", fgColor="2980B9")
    LIGHT2 = PatternFill("solid", fgColor="EBF5FB")

    total_loss = sum(t["realized_profit_usd"] for t in losers)
    avg_loss = total_loss / len(losers) if losers else 0
    worst = losers[0] if losers else None

    summary = [
        ("Token 地址", token_address),
        ("链", chain.upper()),
        ("查询时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("─" * 20, "─" * 30),
        ("Top 100 总地址数", all_count),
        ("亏损地址数", len(losers)),
        ("盈利地址数", all_count - len(losers)),
        ("亏损占比", f"{len(losers)/all_count*100:.1f}%" if all_count else "N/A"),
        ("─" * 20, "─" * 30),
        ("总亏损 (USD)", f"${total_loss:,.2f}"),
        ("平均亏损 (USD)", f"${avg_loss:,.2f}"),
        ("最大单一亏损", f"${worst['realized_profit_usd']:,.2f}" if worst else "N/A"),
        ("最大亏损地址", worst["wallet"] if worst else "N/A"),
    ]

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 50

    for i, (k, v) in enumerate(summary, 1):
        ka = ws2.cell(i, 1, k)
        vb = ws2.cell(i, 2, v)
        if k.startswith("─"):
            continue
        ka.fill = LIGHT2
        vb.fill = LIGHT2
        ka.font = Font(bold=True)

    out = f"worst_traders_{chain}_{token_address[:8]}.xlsx"
    wb.save(out)
    return out


# ──────────────────────────────
# 主函数
# ──────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Worst Traders Scraper — GMGN Top 100 亏损分析")
    parser.add_argument("--token",     required=True,  help="Token 合约地址")
    parser.add_argument("--chain",     default="bsc",  choices=CHAINS, help="链 (默认: bsc)")
    parser.add_argument("--apify-key", default="",     help="Apify API Key (demo 模式可省略)")
    parser.add_argument("--demo",      action="store_true", help="使用模拟数据（无需 API Key）")
    parser.add_argument("--top",       type=int, default=20, help="显示亏损前 N 名（默认: 20）")
    args = parser.parse_args()

    print("\n" + "="*55)
    print("  💸  Worst Traders Scraper  |  GMGN × APIFY")
    print("="*55 + "\n")

    # 获取数据
    if args.demo:
        print("🎭 DEMO 模式 — 使用模拟数据\n")
        raw = demo_data(args.token, args.chain)
    else:
        if not args.apify_key:
            print("❌ 请提供 --apify-key（或使用 --demo 测试）")
            sys.exit(1)
        raw = run_actor(args.token, args.chain, args.apify_key)

    # 解析
    traders = parse_traders(raw)
    total = len(traders)

    # 筛选亏损（realized_profit < 0）
    losers = [t for t in traders if t["realized_profit_usd"] < 0]
    losers.sort(key=lambda x: x["realized_profit_usd"])  # 亏最多在前

    print(f"\n📊 数据摘要:")
    print(f"   总地址: {total}")
    print(f"   亏损:   {len(losers)} ({len(losers)/total*100:.1f}%)")
    print(f"   盈利:   {total - len(losers)} ({(total-len(losers))/total*100:.1f}%)")

    total_loss = sum(t["realized_profit_usd"] for t in losers)
    print(f"\n   总亏损金额:  ${total_loss:,.2f}")
    print(f"   平均亏损:    ${total_loss/len(losers):,.2f}" if losers else "")

    # 控制台输出前 N 名
    print(f"\n💸 亏损最多 Top {min(args.top, len(losers))}:")
    print(f"  {'排名':>4}  {'钱包地址':<44}  {'亏损 (USD)':>12}  {'胜率':>6}")
    print("  " + "-"*72)
    for rank, t in enumerate(losers[:args.top], 1):
        wr = f"{t['win_rate']*100:.0f}%"
        print(f"  {rank:>4}  {t['wallet']:<44}  ${t['realized_profit_usd']:>11,.2f}  {wr:>6}")

    # 导出 Excel
    outfile = export_excel(losers, args.token, args.chain, total)
    print(f"\n✅ Excel 已导出: {outfile}\n")


if __name__ == "__main__":
    main()
